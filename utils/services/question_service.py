from database import get_db
from openpyxl import load_workbook


# -------------------------------
# Upload Questions from Excel
# -------------------------------
def upload_questions(exam_id, excel_file):
    
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        db = get_db()
        cursor = db.cursor()

        # 🔥 DELETE OLD QUESTIONS FIRST
        cursor.execute("""
            DELETE FROM questions
            WHERE exam_id = ?
        """, (exam_id,))

        row_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):

            if not row or all(cell is None for cell in row):
                continue

            if len(row) < 6:
                return False, "Excel format error: Expected 6 columns."

            question, option_a, option_b, option_c, option_d, correct_option = row[:6]

            cursor.execute("""
                INSERT INTO questions (
                    exam_id,
                    question_text,
                    option_a,
                    option_b,
                    option_c,
                    option_d,
                    correct_option
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                exam_id,
                question,
                option_a,
                option_b,
                option_c,
                option_d,
                correct_option.upper()
            ))

            row_count += 1

        db.commit()

        return True, f"{row_count} questions uploaded successfully."

    except Exception as e:
        return False, f"Error uploading questions: {str(e)}"

# -------------------------------
# Get all questions for an exam
# -------------------------------
def get_exam_questions(exam_id):
    """
    Returns all questions for given exam
    """
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT * FROM questions
        WHERE exam_id = ?
        ORDER BY id ASC
    """, (exam_id,))

    return cursor.fetchall()