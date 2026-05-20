from smart_exam_system.extensions import db
from smart_exam_system.models.question import QuestionModel
from openpyxl import load_workbook


# ---------------------------------
# Clean Excel Cell Text
# ---------------------------------

def clean_excel_text(value):

    if value is None:
        return ""

    return str(value).strip()


# ---------------------------------
# Upload Questions
# ---------------------------------

def upload_questions(exam_id, excel_file):

    try:

        workbook = load_workbook(excel_file)
        sheet = workbook.active

        validated_questions = []

        # ---------------------------------
        # VALIDATE FIRST
        # ---------------------------------

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2
        ):

            # Skip completely empty rows
            if not row or all(cell is None for cell in row):
                continue

            if len(row) < 6:
                return (
                    False,
                    f"Excel format error at row {row_index}: Expected 6 columns."
                )

            question, option_a, option_b, option_c, option_d, correct_option = row[:6]

            # Clean values
            question = clean_excel_text(question)
            option_a = clean_excel_text(option_a)
            option_b = clean_excel_text(option_b)
            option_c = clean_excel_text(option_c)
            option_d = clean_excel_text(option_d)

            correct_option = (
                clean_excel_text(correct_option)
                .upper()
            )

            # ---------------------------------
            # Required field validation
            # ---------------------------------

            required_fields = [
                question,
                option_a,
                option_b,
                option_c,
                option_d,
                correct_option
            ]

            if any(field == "" for field in required_fields):

                return (
                    False,
                    f"Missing required data at row {row_index}."
                )

            # ---------------------------------
            # Correct option validation
            # ---------------------------------

            if correct_option not in ["A", "B", "C", "D"]:

                return (
                    False,
                    (
                        f"Invalid correct option at row "
                        f"{row_index}. Use only A, B, C, or D."
                    )
                )

            # Store validated question
            validated_questions.append({
                "question_text": question,
                "option_a": option_a,
                "option_b": option_b,
                "option_c": option_c,
                "option_d": option_d,
                "correct_option": correct_option
            })

        # ---------------------------------
        # No valid questions
        # ---------------------------------

        if not validated_questions:

            return False, "No valid questions found in Excel file."

        # ---------------------------------
        # DELETE OLD QUESTIONS
        # ---------------------------------

        db.session.query(QuestionModel).filter_by(
            exam_id=exam_id
        ).delete(synchronize_session=False)

        # ---------------------------------
        # INSERT NEW QUESTIONS
        # ---------------------------------

        for q in validated_questions:

            new_question = QuestionModel(
                exam_id=exam_id,
                question_text=q["question_text"],
                option_a=q["option_a"],
                option_b=q["option_b"],
                option_c=q["option_c"],
                option_d=q["option_d"],
                correct_option=q["correct_option"]
            )

            db.session.add(new_question)

        db.session.commit()

        return (
            True,
            f"{len(validated_questions)} questions uploaded successfully."
        )

    except Exception as e:

        db.session.rollback()

        return (
            False,
            f"Error uploading questions: {str(e)}"
        )


# ---------------------------------
# Get All Questions For Exam
# ---------------------------------

def get_exam_questions(exam_id):

    return (
        QuestionModel.query
        .filter_by(exam_id=exam_id)
        .order_by(QuestionModel.id.asc())
        .all()
    )